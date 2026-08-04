import asyncio
import tempfile
import unittest
from dataclasses import replace
from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright

from app.main import app
from app.config.settings import Settings
from app.core.dependencies import get_report_service
from app.models.report import (
    CampaignMetrics,
    CampaignRow,
    JobState,
    ReportJob,
    RowResult,
    SheetConnection,
)
from app.services.moengage_browser_service import (
    BrowserAutomationError,
    CIS_BRAND_OPERATOR,
    CIS_BRAND_VALUE,
    CIS_EVENT_BRAND_PATTERN,
    DELIVERY_LOOKBACK_DAYS,
    MoEngageBrowserService,
    PURCHASED_CUSTOMERS_TITLE,
    build_behavior_query_plan,
)
from app.services.moengage_service import MoEngageError, MoEngageService
from app.services.excel_service import ExcelService
from app.services.google_sheet_service import (
    GoogleSheetService,
    previous_completed_week,
)
from app.services.report_service import ReportService
from app.utils.excel_utils import parse_tracking_range


HEADERS = [
    "Date", "Channel", "Brand", "Campaign Name", "Campaign Channel Type",
    "Track Goals for", "Campaign ID", "Campaign Purchased Customers",
    "Campaign Purchased Customers - Online", "Campaign Purchased Customers - Offline",
    "Influenced Revenue", "Influenced Revenue - Online", "Influenced Revenue - Offline",
]


def make_workbook(path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Mastersheet"
    sheet.append(HEADERS)
    sheet.append([
        datetime(2026, 4, 29), "Whatsapp", "Aldo", "Spring offer", "online",
        "29 - 1 May", "readable-123", None, None, None, None, None, None,
    ])
    workbook.save(path)


def campaign_row(channel="WhatsApp", campaign_type="Online", brand="VS"):
    return CampaignRow(
        excel_row=99,
        campaign_date=date(2026, 7, 16),
        brand=brand,
        channel=channel,
        campaign_type=campaign_type,
        campaign_id="example-campaign-id",
        campaign_name="Example campaign",
        tracking_goal="16th July 2026 - 19th July 2026",
        start_date=date(2026, 7, 16),
        end_date=date(2026, 7, 19),
    )


class DateRangeTests(unittest.TestCase):
    def test_previous_completed_week_is_monday_through_sunday(self):
        self.assertEqual(
            previous_completed_week(date(2026, 7, 28)),
            (date(2026, 7, 20), date(2026, 7, 26)),
        )

    def test_cross_month_range_uses_campaign_date(self):
        start, end = parse_tracking_range("29 - 1 May", datetime(2026, 4, 29).date())
        self.assertEqual(start.isoformat(), "2026-04-29")
        self.assertEqual(end.isoformat(), "2026-05-01")

    def test_typo_and_ordinals(self):
        start, end = parse_tracking_range("1st Januay 2026- 3rd Januray 2026", datetime(2026, 1, 1).date())
        self.assertEqual((start.isoformat(), end.isoformat()), ("2026-01-01", "2026-01-03"))


class MoEngageWorkflowTests(unittest.TestCase):
    def test_cis_brand_pm_readonly_check_does_not_mutate_dom(self):
        async def scenario():
            async with async_playwright() as playwright:
                browser = await playwright.chromium.launch(headless=True)
                page = await browser.new_page()
                await page.set_content("""
                    <section id="event">
                      <button data-test="attribute-add-btn">Attributes</button>
                      <div class="mds-attr">
                        <div class="mds-attr__name">Txn_Channel</div>
                        <div class="mds-dropdown">exists</div>
                      </div>
                      <div class="mds-attr">
                        <div class="mds-attr__name">Brand_PM</div>
                        <div class="mds-dropdown">(any of) contains</div>
                        <input value="SP">
                      </div>
                    </section>
                """)
                event = page.locator("#event")
                before = await event.inner_html()
                service = MoEngageBrowserService(
                    Path("profile"), "https://dashboard-03.moengage.com/", {}
                )
                await service._ensure_cis_transaction_brand(page)
                after = await event.inner_html()
                await browser.close()
                self.assertEqual(after, before)

        asyncio.run(scenario())

    def test_browser_mode_reports_brands_from_query_url_map(self):
        with tempfile.TemporaryDirectory() as folder:
            settings = Settings(
                storage_dir=Path(folder),
                moengage_mode="browser",
                moengage_ui_config={
                    "query_url_map": {
                        "VS": "vs-url",
                        "Aldo": "aldo-url",
                        "BBW": "bbw-url",
                    },
                },
            )
            service = MoEngageService(settings)
            self.assertEqual(service.configured_brands(), ["Aldo", "BBW", "VS"])

    def test_login_profile_ids_are_safe_and_case_insensitive(self):
        self.assertEqual(
            MoEngageService.normalize_profile_id("  User.Name+Ops@Example.COM  "),
            "user.name-ops@example.com",
        )
        self.assertEqual(MoEngageService.normalize_profile_id("../../"), "default")

    def test_delivered_event_lookback_is_120_days(self):
        self.assertEqual(DELIVERY_LOOKBACK_DAYS, 120)

    def test_unique_users_sum_daily_values_instead_of_average(self):
        self.assertEqual(
            MoEngageBrowserService._sum_daily_values(["0", "1", "0"]),
            1,
        )

    def test_brand_report_title_variants_are_accepted(self):
        self.assertRegex("Number of Purchased Customers", PURCHASED_CUSTOMERS_TITLE)
        self.assertRegex("Number of Purcahsed Customers", PURCHASED_CUSTOMERS_TITLE)
        self.assertRegex("Campaign Purchased Customers", PURCHASED_CUSTOMERS_TITLE)

    def test_brand_report_title_matches_inside_full_page_content(self):
        page_text = "Dashboard Reporting\nNumber of Purchased Customers\nEvents & filters"
        self.assertIsNotNone(PURCHASED_CUSTOMERS_TITLE.search(page_text))

    def test_brand_specific_behavior_urls_are_selected_case_insensitively(self):
        service = MoEngageBrowserService(
            Path("profile"), "https://dashboard-03.moengage.com/",
            {"query_url_map": {"VS": "vs-url", "Aldo": "aldo-url", "BBW": "bbw-url"}},
        )
        self.assertEqual(service._query_url_for_brand("aldo"), "aldo-url")
        self.assertEqual(service._query_url_for_brand("BBW"), "bbw-url")
        with self.assertRaisesRegex(BrowserAutomationError, "configured for brand 'CK'"):
            service._query_url_for_brand("CK")

    def test_workspace_switch_is_skipped_when_report_dashboard_matches_brand(self):
        query_url = (
            "https://dashboard-03.moengage.com/v4/analytics/v2/behavior"
            "?did=bbw-dashboard&chartId=customers"
        )
        service = MoEngageBrowserService(
            Path("profile"),
            "https://dashboard-03.moengage.com/",
            {
                "query_url_map": {"BBW": query_url},
                "workspace_map": {"BBW": "BBW_IN"},
            },
        )
        asyncio.run(service._switch_workspace(SimpleNamespace(url=query_url), "BBW"))
        self.assertEqual(service.active_workspace, "BBW_IN")

    def test_cis_special_filter_summary_is_recognized(self):
        summary = (
            "Txn_Channel exists AND "
            "Brand_PM (any of) contains SP (case insensitive)"
        )
        self.assertIsNotNone(CIS_EVENT_BRAND_PATTERN.search(summary))
        self.assertEqual(CIS_BRAND_OPERATOR, "(any of) contains")
        self.assertEqual(CIS_BRAND_VALUE, "SP")

    def test_cis_uses_fixed_event_brand_while_agipl_uses_selected_brand(self):
        self.assertTrue(
            MoEngageBrowserService._uses_event_transaction_brand("AGIPL")
        )
        self.assertFalse(
            MoEngageBrowserService._uses_event_transaction_brand("CIS")
        )

    def test_all_provided_dashboard_urls_and_cis_workspace_are_configured(self):
        config = Settings.from_env().moengage_ui_config
        self.assertEqual(
            set(config["query_url_map"]),
            {"VS", "Aldo", "BBW", "CK", "R&B", "BHPC", "Crocs", "AGIPL", "CIS"},
        )
        self.assertEqual(config["workspace_map"]["CIS"], "AL_IN")
        self.assertIn("AGIPL", config["pending_special_workflows"])

    def test_agipl_requires_an_attribution_brand(self):
        service = MoEngageBrowserService(
            Path("profile"), "https://dashboard-03.moengage.com/",
            {"pending_special_workflows": ["AGIPL"]},
        )
        with self.assertRaisesRegex(BrowserAutomationError, "special query logic is still pending"):
            asyncio.run(service._query_recorded_behavior(
                campaign_row(brand="AGIPL"), "unique_users"
            ))

    def test_agipl_brand_names_map_to_transaction_brand_values(self):
        service = MoEngageBrowserService(
            Path("profile"), "https://dashboard-03.moengage.com/",
            {"agipl_attribution_brand_values": {"Aldo": "AL", "CIS": "SP"}},
        )
        self.assertEqual(service._agipl_transaction_brand_value("Aldo"), "AL")
        self.assertEqual(service._agipl_transaction_brand_value("cis"), "SP")
        with self.assertRaisesRegex(BrowserAutomationError, "cannot attribute"):
            service._agipl_transaction_brand_value("AGIPL")

    def test_online_and_offline_transaction_operators(self):
        online = build_behavior_query_plan(campaign_row(campaign_type="Online"), "unique_users")
        offline = build_behavior_query_plan(campaign_row(campaign_type="Offline"), "unique_users")
        self.assertEqual(online.transaction_operator, "exists")
        self.assertEqual(offline.transaction_operator, "does not exist")

    def test_delivery_event_matches_channel(self):
        expected = {
            "WhatsApp": "WhatsApp Message Delivered",
            "SMS": "SMS Delivered",
            "RCS": "RCS Delivered",
        }
        for channel, event in expected.items():
            with self.subTest(channel=channel):
                plan = build_behavior_query_plan(campaign_row(channel=channel), "unique_users")
                self.assertEqual(plan.delivery_event, event)

    def test_revenue_uses_sum_of_order_net_value(self):
        plan = build_behavior_query_plan(campaign_row(), "total_revenue")
        self.assertEqual(plan.analysis_type, "Aggregation")
        self.assertEqual(plan.aggregation, "Sum")
        self.assertEqual(plan.aggregation_attribute, "Order_Net_Val")

    def test_overall_is_not_sent_as_a_single_query(self):
        with self.assertRaises(BrowserAutomationError):
            build_behavior_query_plan(campaign_row(campaign_type="Overall"), "unique_users")

    def test_overall_sums_online_and_offline_metrics(self):
        with tempfile.TemporaryDirectory() as folder:
            settings = Settings(storage_dir=Path(folder), moengage_mode="browser")
            service = MoEngageService(settings)
            service.browser.query_metric = AsyncMock(side_effect=[75, 10, 247287, 12000])
            result = asyncio.run(service.fetch_metrics(campaign_row(campaign_type="Overall")))
        self.assertEqual(result.unique_users, 85)
        self.assertEqual(result.total_revenue, 259287)
        self.assertEqual(result.online_unique_users, 75)
        self.assertEqual(result.offline_unique_users, 10)
        self.assertEqual(result.online_revenue, 247287)
        self.assertEqual(result.offline_revenue, 12000)
        queried_types = [call.args[0].campaign_type for call in service.browser.query_metric.await_args_list]
        self.assertEqual(queried_types, ["Online", "Offline", "Online", "Offline"])

    def test_nonzero_revenue_with_zero_unique_users_is_rejected(self):
        with tempfile.TemporaryDirectory() as folder:
            settings = Settings(storage_dir=Path(folder), moengage_mode="browser")
            service = MoEngageService(settings)
            service.browser.query_metric = AsyncMock(side_effect=[0, 7498])
            with self.assertRaisesRegex(
                MoEngageError, "unique_users=0, revenue=7498.0"
            ):
                asyncio.run(service.fetch_metrics(campaign_row()))

    def test_nonzero_unique_users_with_zero_revenue_is_rejected(self):
        with tempfile.TemporaryDirectory() as folder:
            settings = Settings(storage_dir=Path(folder), moengage_mode="browser")
            service = MoEngageService(settings)
            service.browser.query_metric = AsyncMock(side_effect=[4, 0])
            with self.assertRaisesRegex(
                MoEngageError, "unique_users=4, revenue=0.0"
            ):
                asyncio.run(service.fetch_metrics(campaign_row()))

    def test_zero_users_and_zero_revenue_are_consistent(self):
        metrics = MoEngageService._typed_metrics("Offline", 0, 0)
        self.assertEqual(metrics.unique_users, 0)
        self.assertEqual(metrics.total_revenue, 0)


class ExcelTests(unittest.TestCase):
    def test_blank_campaign_input_is_reported_with_row_and_column(self):
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "input.xlsx"
            make_workbook(source)
            workbook = load_workbook(source)
            sheet = workbook["Mastersheet"]
            sheet.append([
                datetime(2026, 4, 30), "Whatsapp", "Aldo", "Missing ID", "online",
                "30 April - 1 May", None, None, None, None, None, None, None,
            ])
            workbook.save(source)
            _, rows, warnings = ExcelService().read_campaigns(source)
        self.assertEqual(len(rows), 1)
        self.assertIn("Row 3: blank required cell(s): Campaign ID", warnings)

    def test_online_metrics_go_to_online_columns(self):
        with tempfile.TemporaryDirectory() as folder:
            source, output = Path(folder) / "input.xlsx", Path(folder) / "output.xlsx"
            make_workbook(source)
            service = ExcelService()
            _, rows, warnings = service.read_campaigns(source)
            self.assertFalse(warnings)
            from app.models.report import CampaignMetrics
            service.write_metrics(source, output, [(rows[0], CampaignMetrics(12, 345.5))])
            sheet = load_workbook(output, data_only=True)["Mastersheet"]
            self.assertEqual(sheet["H2"].value, 12)
            self.assertEqual(sheet["I2"].value, 12)
            self.assertEqual(sheet["K2"].value, 345.5)
            self.assertEqual(sheet["L2"].value, 345.5)

    def test_offline_metrics_go_to_total_and_offline_columns(self):
        with tempfile.TemporaryDirectory() as folder:
            source, output = Path(folder) / "input.xlsx", Path(folder) / "output.xlsx"
            make_workbook(source)
            service = ExcelService()
            _, rows, _ = service.read_campaigns(source)
            offline = replace(rows[0], campaign_type="Offline")
            from app.models.report import CampaignMetrics
            service.write_metrics(source, output, [(offline, CampaignMetrics(9, 250.0))])
            sheet = load_workbook(output, data_only=True)["Mastersheet"]
            self.assertEqual(sheet["H2"].value, 9)
            self.assertEqual(sheet["J2"].value, 9)
            self.assertEqual(sheet["K2"].value, 250)
            self.assertEqual(sheet["M2"].value, 250)

    def test_overall_metrics_write_breakdown_and_summed_totals(self):
        with tempfile.TemporaryDirectory() as folder:
            source, output = Path(folder) / "input.xlsx", Path(folder) / "output.xlsx"
            make_workbook(source)
            service = ExcelService()
            _, rows, _ = service.read_campaigns(source)
            overall = replace(rows[0], campaign_type="Overall")
            from app.models.report import CampaignMetrics
            metrics = CampaignMetrics(
                unique_users=115,
                total_revenue=292463,
                online_unique_users=100,
                offline_unique_users=15,
                online_revenue=250000,
                offline_revenue=42463,
            )
            service.write_metrics(source, output, [(overall, metrics)])
            sheet = load_workbook(output, data_only=True)["Mastersheet"]
            self.assertEqual(
                [sheet.cell(2, column).value for column in range(8, 14)],
                [115, 100, 15, 292463, 250000, 42463],
            )


class GoogleSheetWriterTests(unittest.TestCase):
    def test_blank_warnings_are_limited_to_the_selected_week(self):
        service = GoogleSheetService(Path("service-account.json"))
        worksheet = MagicMock()
        worksheet.get.return_value = [
            HEADERS,
            [
                "19/07/2026", "Whatsapp", "Aldo", "Older missing ID", "online",
                "19 - 20 July", "", "", "", "", "", "", "",
            ],
            [
                "20/07/2026", "Whatsapp", "Aldo", "Weekly missing ID", "online",
                "20 - 21 July", "", "", "", "", "", "", "",
            ],
            [
                "26/07/2026", "Whatsapp", "Aldo", "Weekly valid", "online",
                "26 - 27 July", "readable-126", "", "", "", "", "", "",
            ],
            [
                "27/07/2026", "Whatsapp", "Aldo", "Current week missing ID", "online",
                "27 - 28 July", "", "", "", "", "", "", "",
            ],
        ]
        rows, warnings = service._read_worksheet(
            worksheet, date(2026, 7, 20), date(2026, 7, 26)
        )
        self.assertEqual([row.campaign_id for row in rows], ["readable-126"])
        self.assertEqual(
            warnings,
            ["Row 3: blank required cell(s): Campaign ID"],
        )

    def test_blank_google_sheet_input_is_reported_instead_of_silently_skipped(self):
        service = GoogleSheetService(Path("service-account.json"))
        worksheet = MagicMock()
        worksheet.get.return_value = [
            HEADERS,
            [
                "29/04/2026", "Whatsapp", "Aldo", "Spring offer", "online",
                "29 - 1 May", "readable-123", "", "", "", "", "", "",
            ],
            [
                "30/04/2026", "Whatsapp", "Aldo", "Missing ID", "online",
                "30 April - 1 May", "", "", "", "", "", "", "",
            ],
        ]
        rows, warnings = service._read_worksheet(worksheet)
        self.assertEqual(len(rows), 1)
        self.assertIn("Row 3: blank required cell(s): Campaign ID", warnings)

    def test_overall_writes_exact_aa_through_af_breakdown(self):
        with tempfile.TemporaryDirectory() as folder:
            service = GoogleSheetService(Path(folder) / "service-account.json")
            worksheet = MagicMock()
            client = MagicMock()
            client.open_by_key.return_value.worksheet.return_value = worksheet
            service._client = MagicMock(return_value=client)
            metrics = CampaignMetrics(
                unique_users=115,
                total_revenue=292463,
                online_unique_users=100,
                offline_unique_users=15,
                online_revenue=250000,
                offline_revenue=42463,
            )
            asyncio.run(service.write_metrics(
                "sheet-id", "Mastersheet",
                campaign_row(campaign_type="Overall"), metrics,
            ))
            updates = worksheet.batch_update.call_args.args[0]
            self.assertEqual(
                {item["range"]: item["values"][0][0] for item in updates},
                {
                    "AA99": 115,
                    "AB99": 100,
                    "AC99": 15,
                    "AD99": 292463,
                    "AE99": 250000,
                    "AF99": 42463,
                },
            )


class ReportFilterTests(unittest.TestCase):
    def test_brand_filter_is_case_insensitive_and_excludes_other_brands(self):
        rows = [campaign_row(brand="VS"), campaign_row(brand="Aldo")]
        filtered = ReportService._filter_brands(rows, ["vs"])
        self.assertEqual([row.brand for row in filtered], ["VS"])

    def test_combined_filters_match_brand_channel_and_sent_date(self):
        rows = [
            campaign_row(channel="WhatsApp", brand="VS"),
            campaign_row(channel="SMS", brand="VS"),
            replace(campaign_row(channel="WhatsApp", brand="VS"), campaign_date=date(2026, 7, 17)),
            campaign_row(channel="WhatsApp", brand="Aldo"),
        ]
        filtered = ReportService._filter_campaigns(
            rows, brands=["vs"], channels=["whatsapp"], sent_date=date(2026, 7, 16)
        )
        self.assertEqual(len(filtered), 1)
        self.assertEqual(filtered[0].channel, "WhatsApp")

    def test_empty_filter_lists_preserve_all_campaigns(self):
        rows = [campaign_row(channel="WhatsApp"), campaign_row(channel="SMS")]
        self.assertEqual(ReportService._filter_campaigns(rows), rows)

    def test_sent_date_range_is_inclusive(self):
        rows = [
            replace(campaign_row(), campaign_date=date(2026, 7, 14)),
            replace(campaign_row(), campaign_date=date(2026, 7, 15)),
            replace(campaign_row(), campaign_date=date(2026, 7, 17)),
            replace(campaign_row(), campaign_date=date(2026, 7, 18)),
        ]
        filtered = ReportService._filter_campaigns(
            rows,
            sent_date_from=date(2026, 7, 15),
            sent_date_to=date(2026, 7, 17),
        )
        self.assertEqual(
            [row.campaign_date for row in filtered],
            [date(2026, 7, 15), date(2026, 7, 17)],
        )


class FailedRetryTests(unittest.TestCase):
    def test_retry_job_contains_only_failed_sheet_rows(self):
        async def scenario():
            with tempfile.TemporaryDirectory() as folder:
                service = ReportService(Settings(storage_dir=Path(folder)))
                connection = SheetConnection(
                    id="sheet-1", spreadsheet_id="spreadsheet", spreadsheet_url="url",
                    spreadsheet_title="Master", worksheet_title="Mastersheet", row_count=2,
                    brands=["BBW"], channels=["WhatsApp"], campaign_types=["Online"],
                    sent_dates=[date(2026, 7, 15)], preview=[],
                )
                original = ReportJob(
                    id="original", upload_id=connection.id, filename="Master",
                    state=JobState.COMPLETED,
                    agipl_attribution_brand="Aldo",
                )
                original.results.extend([
                    RowResult(
                        excel_row=2310, brand="BBW", channel="WhatsApp",
                        campaign_type="Online", campaign_id="failed", campaign_name="Failed",
                        date_range="range", status="failed",
                    ),
                    RowResult(
                        excel_row=2311, brand="BBW", channel="WhatsApp",
                        campaign_type="Offline", campaign_id="success", campaign_name="Success",
                        date_range="range", status="success",
                    ),
                ])
                service.sheet_connections[connection.id] = connection
                service.jobs[original.id] = original
                service._process_sheet = AsyncMock()
                retry = service.retry_failed_sheet_job(original.id)
                await service.tasks[retry.id]
                self.assertEqual(retry.upload_id, connection.id)
                self.assertEqual(retry.agipl_attribution_brand, "Aldo")
                self.assertEqual(
                    service._process_sheet.await_args.kwargs["row_numbers"], {2310}
                )
                self.assertEqual(
                    service._process_sheet.await_args.kwargs["agipl_attribution_brand"], "Aldo"
                )

        asyncio.run(scenario())


class ApiTests(unittest.TestCase):
    def test_sheet_connection_exposes_previous_week_problem_window(self):
        service = get_report_service()
        original_connect_sheet = service.connect_sheet
        service.connect_sheet = AsyncMock(return_value=SheetConnection(
            id="weekly-sheet",
            spreadsheet_id="spreadsheet",
            spreadsheet_url="https://docs.google.com/spreadsheets/d/test",
            spreadsheet_title="Master",
            worksheet_title="Mastersheet",
            row_count=1,
            brands=["VS"],
            channels=["WhatsApp"],
            campaign_types=["Online"],
            sent_dates=[date(2026, 7, 20)],
            preview=[],
            warning_sent_date_from=date(2026, 7, 20),
            warning_sent_date_to=date(2026, 7, 26),
        ))
        try:
            with TestClient(app) as client:
                response = client.post("/api/google/connect", json={
                    "spreadsheet_url": "https://docs.google.com/spreadsheets/d/test",
                    "worksheet_name": "Mastersheet",
                })
        finally:
            service.connect_sheet = original_connect_sheet
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["warning_sent_date_from"], "2026-07-20")
        self.assertEqual(response.json()["warning_sent_date_to"], "2026-07-26")

    def test_agipl_job_requires_attribution_brand(self):
        with TestClient(app) as client:
            response = client.post("/api/jobs", json={
                "sheet_connection_id": "missing",
                "brands": ["AGIPL"],
                "channels": ["WhatsApp"],
                "sent_date_from": "2026-07-15",
                "sent_date_to": "2026-07-16",
            })
        self.assertEqual(response.status_code, 422)
        self.assertIn("attribution brand", response.text)

    def test_health_exposes_live_connection_state(self):
        with TestClient(app) as client:
            response = client.get("/api/health")
            self.assertEqual(response.status_code, 200)
            self.assertIn("google_configured", response.json())
            self.assertIn("moengage_connected", response.json())

    def test_mock_mode_cannot_write_fake_metrics_by_default(self):
        mock_settings = SimpleNamespace(
            moengage_mode="mock",
            allow_mock_writes=False,
        )
        with patch("app.api.routes.settings", mock_settings), TestClient(app) as client:
            response = client.post("/api/jobs", json={"sheet_connection_id": "sheet"})
        self.assertEqual(response.status_code, 409)
        self.assertIn("Mock metrics are disabled", response.json()["detail"])

    def test_second_job_is_rejected_while_automation_is_running(self):
        service = get_report_service()
        active = ReportJob(id="already-running", upload_id="sheet", filename="Master Sheet")
        active.state = JobState.PROCESSING
        service.jobs[active.id] = active
        try:
            with TestClient(app) as client:
                response = client.post("/api/jobs", json={"sheet_connection_id": "sheet"})
            self.assertEqual(response.status_code, 409)
            self.assertIn("already running", response.json()["detail"])
        finally:
            service.jobs.pop(active.id, None)

    def test_invalid_google_key_is_rejected_without_installing_it(self):
        with TestClient(app) as client:
            response = client.post(
                "/api/google/credentials",
                files={"credential": ("invalid.json", b"{}", "application/json")},
            )
        self.assertEqual(response.status_code, 422)
        self.assertIn("service-account", response.json()["detail"])

    def test_complete_job_results_can_be_downloaded_as_csv(self):
        service = get_report_service()
        job = ReportJob(id="csv-test", upload_id="sheet", filename="Master Sheet")
        job.state = JobState.COMPLETED
        job.results.append(RowResult(
            excel_row=22, brand="VS", channel="WhatsApp", campaign_type="Online",
            campaign_id="campaign-1", campaign_name="Campaign One",
            date_range="2026-07-16 → 2026-07-19", status="success",
            unique_users=12, total_revenue=3456,
            online_unique_users=12, online_revenue=3456,
        ))
        service.jobs[job.id] = job
        with TestClient(app) as client:
            response = client.get("/api/jobs/csv-test/results.csv")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Campaign One", response.text)
        self.assertIn("Unique Users - Online", response.text)
        self.assertIn("Revenue - Offline", response.text)
        self.assertIn("attachment", response.headers["content-disposition"])
        service.jobs.pop(job.id, None)


if __name__ == "__main__":
    unittest.main()
