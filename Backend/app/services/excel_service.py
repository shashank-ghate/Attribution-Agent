from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.models.report import CampaignMetrics, CampaignRow
from app.utils.excel_utils import (
    normalize_campaign_type,
    normalize_channel,
    parse_tracking_range,
)


class WorkbookValidationError(ValueError):
    pass


class ExcelService:
    INPUT_HEADERS = (
        "Date",
        "Channel",
        "Brand",
        "Campaign Name",
        "Campaign Channel Type",
        "Track Goals for",
        "Campaign ID",
    )
    REQUIRED_HEADERS = {
        "Date",
        "Channel",
        "Brand",
        "Campaign Name",
        "Campaign Channel Type",
        "Track Goals for",
        "Campaign ID",
        "Campaign Purchased Customers",
        "Campaign Purchased Customers - Online",
        "Campaign Purchased Customers - Offline",
        "Influenced Revenue",
        "Influenced Revenue - Online",
        "Influenced Revenue - Offline",
    }
    TARGETS = {
        "Overall": ("Campaign Purchased Customers", "Influenced Revenue"),
        "Online": ("Campaign Purchased Customers - Online", "Influenced Revenue - Online"),
        "Offline": ("Campaign Purchased Customers - Offline", "Influenced Revenue - Offline"),
    }

    @staticmethod
    def metric_values(campaign_type: str, metrics: CampaignMetrics) -> dict[str, int | float]:
        """Map totals and channel-type breakdowns to their sheet headers."""
        values: dict[str, int | float] = {
            "Campaign Purchased Customers": metrics.unique_users,
            "Influenced Revenue": metrics.total_revenue,
        }
        if campaign_type == "Overall":
            breakdown = (
                metrics.online_unique_users,
                metrics.offline_unique_users,
                metrics.online_revenue,
                metrics.offline_revenue,
            )
            if any(value is None for value in breakdown):
                raise WorkbookValidationError(
                    "Overall campaign metrics must include online and offline breakdowns"
                )
            values.update({
                "Campaign Purchased Customers - Online": metrics.online_unique_users,
                "Campaign Purchased Customers - Offline": metrics.offline_unique_users,
                "Influenced Revenue - Online": metrics.online_revenue,
                "Influenced Revenue - Offline": metrics.offline_revenue,
            })
        elif campaign_type == "Online":
            values.update({
                "Campaign Purchased Customers - Online": (
                    metrics.online_unique_users
                    if metrics.online_unique_users is not None else metrics.unique_users
                ),
                "Influenced Revenue - Online": (
                    metrics.online_revenue
                    if metrics.online_revenue is not None else metrics.total_revenue
                ),
            })
        elif campaign_type == "Offline":
            values.update({
                "Campaign Purchased Customers - Offline": (
                    metrics.offline_unique_users
                    if metrics.offline_unique_users is not None else metrics.unique_users
                ),
                "Influenced Revenue - Offline": (
                    metrics.offline_revenue
                    if metrics.offline_revenue is not None else metrics.total_revenue
                ),
            })
        else:
            raise WorkbookValidationError(f"Unsupported campaign type {campaign_type!r}")
        return values

    def _open(self, path: Path, data_only: bool = False):
        try:
            return load_workbook(path, data_only=data_only, keep_links=True)
        except Exception as exc:
            raise WorkbookValidationError(f"Could not open Excel workbook: {exc}") from exc

    def _sheet_and_headers(self, workbook):
        sheet = workbook["Mastersheet"] if "Mastersheet" in workbook.sheetnames else workbook.active
        headers = {
            str(cell.value).strip(): cell.column
            for cell in sheet[1]
            if cell.value is not None and str(cell.value).strip()
        }
        missing = sorted(self.REQUIRED_HEADERS - set(headers))
        if missing:
            raise WorkbookValidationError("Missing required columns: " + ", ".join(missing))
        return sheet, headers

    def read_campaigns(self, path: Path) -> tuple[str, list[CampaignRow], list[str]]:
        workbook = self._open(path, data_only=True)
        sheet, headers = self._sheet_and_headers(workbook)
        rows: list[CampaignRow] = []
        warnings: list[str] = []
        for row_number in range(2, sheet.max_row + 1):
            input_values = {
                header: sheet.cell(row_number, headers[header]).value
                for header in self.INPUT_HEADERS
            }
            if not any(
                value is not None and str(value).strip()
                for value in input_values.values()
            ):
                continue
            blank_headers = [
                header
                for header, value in input_values.items()
                if value is None or not str(value).strip()
            ]
            if blank_headers:
                warnings.append(
                    f"Row {row_number}: blank required cell(s): "
                    + ", ".join(blank_headers)
                )
                continue
            campaign_id = input_values["Campaign ID"]
            try:
                raw_date = input_values["Date"]
                campaign_date = raw_date.date() if hasattr(raw_date, "date") else raw_date
                brand = str(input_values["Brand"]).strip()
                channel = normalize_channel(input_values["Channel"])
                campaign_type = normalize_campaign_type(
                    input_values["Campaign Channel Type"]
                )
                tracking_goal = input_values["Track Goals for"]
                start_date, end_date = parse_tracking_range(tracking_goal, campaign_date)
                metric_value = lambda header: sheet.cell(row_number, headers[header]).value
                rows.append(
                    CampaignRow(
                        excel_row=row_number,
                        campaign_date=campaign_date,
                        brand=brand,
                        channel=channel,
                        campaign_type=campaign_type,
                        campaign_id=str(campaign_id).strip(),
                        campaign_name=str(sheet.cell(row_number, headers["Campaign Name"]).value or "").strip(),
                        tracking_goal=str(tracking_goal or ""),
                        start_date=start_date,
                        end_date=end_date,
                        existing_unique_users=metric_value("Campaign Purchased Customers"),
                        existing_revenue=metric_value("Influenced Revenue"),
                        existing_online_unique_users=metric_value("Campaign Purchased Customers - Online"),
                        existing_offline_unique_users=metric_value("Campaign Purchased Customers - Offline"),
                        existing_online_revenue=metric_value("Influenced Revenue - Online"),
                        existing_offline_revenue=metric_value("Influenced Revenue - Offline"),
                    )
                )
            except (ValueError, TypeError) as exc:
                warnings.append(f"Row {row_number}: {exc}")
        if not rows:
            raise WorkbookValidationError("No processable campaign rows were found")
        return sheet.title, rows, warnings

    def write_metrics(
        self,
        source_path: Path,
        output_path: Path,
        updates: Iterable[tuple[CampaignRow, CampaignMetrics]],
    ) -> None:
        workbook = self._open(source_path, data_only=False)
        sheet, headers = self._sheet_and_headers(workbook)
        success_fill = PatternFill(fill_type="solid", fgColor="E8F7F0")
        for campaign, metrics in updates:
            for header, value in self.metric_values(campaign.campaign_type, metrics).items():
                cell = sheet.cell(campaign.excel_row, headers[header])
                cell.value = value
                cell.number_format = '#,##0.00' if "Revenue" in header else "0"
                cell.fill = copy(success_fill)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
